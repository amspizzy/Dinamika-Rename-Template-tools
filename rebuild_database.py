"""
Rebuild database.xlsx dari official price list files di ~/Downloads/database/
- Baca semua file price list resmi per kategori
- Update/insert SKU dengan data akurat (Brand, Category, Size, Folder)
- Item yang tidak ada di price list tetap dipertahankan
"""

import re, os, sys, glob
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database.xlsx")
PRICE_DIR = os.path.expanduser("~/Downloads/database")

import import_items as ii

FILE_CONFIG = [
    {
        "file": "PRICE LIST GLASSWARE DINAMIKA 2026.xlsx",
        "cat": "Glassware",
        "header_row": 3,
        "data_start": 4,
        "col_map": {"code": 2, "desc": 3, "brand": 4, "subcat": None},
    },
    {
        "file": "2. PRICE LIST PT CMSI 2026-TISSUE CULTURE.xlsx",
        "cat": "Tissue Culture",
        "header_row": 3,
        "data_start": 4,
        "col_map": {"code": 2, "desc": 3, "brand": 6, "subcat": None},
    },
    {
        "file": "3. PRICE LIST PT CMSI 2026-PLASTIC WARE.xlsx",
        "cat": "Plastic Ware",
        "header_row": 3,
        "data_start": 4,
        "col_map": {"code": 2, "desc": 3, "brand": 4, "subcat": None},
    },
    {
        "file": "4. PRICE LIST PT CMSI 2026-PORCELAIN WARE.xlsx",
        "cat": "Porcelain Ware",
        "header_row": 3,
        "data_start": 4,
        "col_map": {"code": 2, "desc": 3, "brand": 4, "subcat": None},
    },
    {
        "file": "5. PRICE LIST PT CMSI 2026-TOOLS.xlsx",
        "cat": "Tools",
        "header_row": 3,
        "data_start": 4,
        "col_map": {"code": 2, "desc": 3, "brand": 4, "subcat": None},
    },
    {
        "file": "7. PRICE LIST PT CMSI 2026-EQUIPMENT.xlsx",
        "cat": "Equipment",
        "header_row": 4,
        "data_start": 5,
        "col_map": {"code": 3, "desc": 4, "brand": 5, "subcat": 2},
    },
    {
        "file": "List Chemical.xlsx",
        "cat": "Chemicals",
        "header_row": 1,
        "data_start": 2,
        "col_map": {"code": 1, "desc": 2, "brand": None, "subcat": None},
    },
]


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def extract_size_from_sku(sku, desc):
    """Extract size from chemical SKU format 1.XXXXX.CCCC or 8.XXXXX.CCCC or -100G/-500G suffix"""
    parts = sku.split(".")
    if len(parts) >= 3:
        last = re.sub(r"[^0-9]", "", parts[-1])
        if len(last) == 4 and last.isdigit():
            val = int(last)
            if val > 0:
                return f"{val} ml"
    suffix = re.search(r"-(\d+)\s*(G|ML|L|KG|MG)\b", sku, re.IGNORECASE)
    if suffix:
        val = int(suffix.group(1))
        unit = suffix.group(2).upper()
        if unit == "G":
            return f"{val} g"
        elif unit == "KG":
            return f"{val} kg"
        elif unit == "ML":
            return f"{val} ml"
        elif unit == "L":
            return f"{val} L"
        elif unit == "MG":
            return f"{val} mg"
    return ii.extract_size(desc)


def detect_subcategory_from_desc(desc, main_cat):
    text_lower = desc.lower()
    for subcat, pattern in ii.SUBCATEGORY_KEYWORDS:
        if re.search(pattern, text_lower):
            return subcat
    first_part = desc.split(",")[0].strip()
    first_part = re.sub(r"\(.*?\)", "", first_part).strip()
    words = re.findall(r"[A-Za-z]+", first_part)
    meaningful = [w for w in words if len(w) > 2 and w.lower() not in ("the", "for", "with", "and")]
    if meaningful:
        return meaningful[0].title()
    return main_cat


def generate_vision_keywords(main_cat, subcat):
    templates = {
        "Glassware": "clear glass laboratory {s} equipment",
        "Tissue Culture": "sterile plastic laboratory {s} for cell culture",
        "Plastic Ware": "plastic laboratory {s} disposable equipment",
        "Porcelain Ware": "white porcelain laboratory {s} heat resistant",
        "Tools": "laboratory {s} tool stainless steel",
        "Filter Paper": "laboratory grade {s} for filtration",
        "Equipment": "laboratory {s} equipment",
        "Chemicals": "laboratory grade {s} chemical reagent",
    }
    tpl = templates.get(main_cat, "laboratory {s}")
    return tpl.format(s=subcat.lower())


def read_price_list(config):
    path = os.path.join(PRICE_DIR, config["file"])
    if not os.path.exists(path):
        print(f"  FILE TIDAK DITEMUKAN: {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cm = config["col_map"]

    items = []
    for r in range(config["data_start"], ws.max_row + 1):
        code = ws.cell(r, cm["code"]).value
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue

        desc = str(ws.cell(r, cm["desc"]).value or "").strip()
        if not desc:
            continue

        brand_raw = ws.cell(r, cm["brand"]).value if cm["brand"] else None
        brand = str(brand_raw).strip() if brand_raw else ""

        subcat_raw = ws.cell(r, cm["subcat"]).value if cm["subcat"] else None
        subcat = str(subcat_raw).strip() if subcat_raw else ""

        items.append({"sku": code, "desc": desc, "brand": brand, "subcat": subcat})

    return items


def build_official_data():
    print("=== Baca official price list files ===\n")
    all_items = {}
    total = 0

    for fc in FILE_CONFIG:
        items = read_price_list(fc)
        total += len(items)
        for item in items:
            sku = item["sku"]
            item["main_cat"] = fc["cat"]
            all_items[sku] = item
        print(f"  {fc['cat']:20s}: {len(items):5d} items")

    print(f"\nTotal official items: {total}")
    return all_items, total


def main():
    official, total_official = build_official_data()

    print("\n=== Update database ===\n")

    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Price List"]

    updated = 0
    added = 0
    kept = 0

    existing_skus = {}
    rows_to_keep = []

    for r in range(3, ws.max_row + 1):
        sku = str(ws.cell(r, 1).value or "").strip()
        if not sku:
            continue
        existing_skus[sku] = r
        rows_to_keep.append(r)

    official_added = 0

    for r in range(3, ws.max_row + 1):
        sku = str(ws.cell(r, 1).value or "").strip()
        if not sku:
            continue

        if sku in official:
            off = official[sku]
            desc = off["desc"]
            brand = off["brand"]
            main_cat = off["main_cat"]
            subcat = off["subcat"] if off["subcat"] else detect_subcategory_from_desc(desc, main_cat)
            size = extract_size_from_sku(sku, desc) if main_cat == "Chemicals" else ii.extract_size(desc)
            ocr_kw = ii.generate_ocr_keywords(desc, brand, size, subcat)
            vision_kw = generate_vision_keywords(main_cat, subcat)
            folder = f"{main_cat.replace(' ', '_')}/{slugify(subcat)}"

            ws.cell(r, 2, desc)
            ws.cell(r, 3, brand if brand else "")
            ws.cell(r, 4, subcat)
            ws.cell(r, 5, size)
            ws.cell(r, 6, ocr_kw)
            ws.cell(r, 7, vision_kw)
            ws.cell(r, 8, subcat)
            ws.cell(r, 9, folder)
            ws.cell(r, 10, "=TRUE()")
            updated += 1
        else:
            kept += 1

    next_row = ws.max_row + 1
    for sku, off in official.items():
        if sku not in existing_skus:
            desc = off["desc"]
            brand = off["brand"]
            main_cat = off["main_cat"]
            subcat = off["subcat"] if off["subcat"] else detect_subcategory_from_desc(desc, main_cat)
            size = extract_size_from_sku(sku, desc) if main_cat == "Chemicals" else ii.extract_size(desc)
            ocr_kw = ii.generate_ocr_keywords(desc, brand, size, subcat)
            vision_kw = generate_vision_keywords(main_cat, subcat)
            folder = f"{main_cat.replace(' ', '_')}/{slugify(subcat)}"

            ws.cell(next_row, 1, sku)
            ws.cell(next_row, 2, desc)
            ws.cell(next_row, 3, brand if brand else "")
            ws.cell(next_row, 4, subcat)
            ws.cell(next_row, 5, size)
            ws.cell(next_row, 6, ocr_kw)
            ws.cell(next_row, 7, vision_kw)
            ws.cell(next_row, 8, subcat)
            ws.cell(next_row, 9, folder)
            ws.cell(next_row, 10, "=TRUE()")
            next_row += 1
            added += 1

    size_fixed = 0
    for r in range(3, next_row):
        curr_size = ws.cell(r, 5).value
        if not curr_size or str(curr_size).strip() in ("", "None"):
            sku = str(ws.cell(r, 1).value or "").strip()
            desc = str(ws.cell(r, 2).value or "").strip()
            folder = str(ws.cell(r, 9).value or "")
            mc = folder.split("/")[0].replace("_", " ")
            if not desc:
                continue
            new_size = extract_size_from_sku(sku, desc) if mc == "Chemicals" else ii.extract_size(desc)
            if new_size:
                ws.cell(r, 5, new_size)
                ocr_kw = ii.generate_ocr_keywords(desc, str(ws.cell(r, 3).value or ""), new_size, str(ws.cell(r, 4).value or ""))
                ws.cell(r, 6, ocr_kw)
                size_fixed += 1

    wb.save(DB_PATH)

    total_final = next_row - 1 - (ws.max_row - total_official)
    print(f"Updated: {updated}")
    print(f"Added (new from official): {added}")
    print(f"Size fixed (post-process): {size_fixed}")
    print(f"Kept (not in official): {kept}")
    print(f"Total official items processed: {total_official}")
    print(f"\nDatabase saved. Total rows: {next_row - 1}")

    from collections import Counter
    folders = Counter()
    for r in range(3, next_row):
        f = str(ws.cell(r, 9).value or "")
        mc = f.split("/")[0] if "/" in f else f
        folders[mc] += 1

    print("\nBreakdown per kategori:")
    for k in ["Glassware", "Tissue_Culture", "Plastic_Ware", "Porcelain_Ware", "Tools", "Filter_Paper", "Equipment", "Chemicals"]:
        print(f"  {k:20s}: {folders.get(k, 0):6d}")


if __name__ == "__main__":
    main()
