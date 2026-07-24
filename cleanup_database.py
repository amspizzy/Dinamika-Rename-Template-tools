"""
Perbaiki kategori & isi size kosong di database.xlsx
- SKU 1.xxxx / 8.xxxx → Chemicals (kecuali yang beneran Equipment)
- Isi size dari 4 digit terakhir SKU untuk chemicals
- Update Output Folder sesuai kategori baru
"""

import re, os
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database.xlsx")

EQUIPMENT_KEYWORDS = re.compile(
    r"digital|electronic|balance|ph[- ]?meter|conductivity|turbidity|"
    r"thermoreactor|spectroquant|oven|incubator|autoclave|centrifuge|"
    r"shaker|stirrer|hotplate|hot.?plate|water.?bath|furnace|pump|"
    r"mixer|homogenizer|vortex|rotary|laminar|hood|heater|sonicator|"
    r"microscope|refracto|polarimeter|colorimeter|photometer|"
    r"prove\s\d+|halogen.?lamp|pipecheck|combicheck|photocheck|"
    r"fixture\s|housing\s|alignment.?tool|bio.?convex.?lens|"
    r"code.?reader|move\sdc|cultura.?mini|multirack|"
    r"foot.?pump|electronic.?board", re.I
)

SIZE_UNITS = {"1 L": 1000, "2.5 L": 2500, "500 ML": 500, "100 ML": 100,
              "250 ML": 250, "25 ML": 25, "10 ML": 10, "5 ML": 5,
              "50 ML": 50, "4 L": 4000, "5 L": 5000, "1000 ML": 1000}

CATEGORIES_AMAN = {
    "Glassware", "Tissue_Culture", "Plastic_Ware", "Porcelain_Ware",
    "Tools", "Filter_Paper",
}

STAY_EQUIPMENT = re.compile(
    r"(?:test\s+(?:kit|strip|set|reagent|indicator|solution|"
    r"tablet|disc|bottle|pack|box|equipment)|"
    r"reagent\s+(?:set|kit|test|for|solution|strip|disc|tablet)|"
    r"indicator\s+(?:solution|strip|test|paper|set|reagent|tablet)|"
    r"standard\s+(?:solution|set|kit|for)|"
    r"buffer\s+(?:solution|tablet|powder|sachet|for)|"
    r"control\s+(?:set|kit|solution|standard|serum|for)|"
    r"medium\s+(?:for|culture|agar|broth|selective|differential)|"
    r"agar\s+(?:for|plating|slant|deep|nutrient|selective)|"
    r"broth\s+(?:for|culture|enrichment|selective)|"
    r"antigen|antibody|antiserum|antitoxin|toxin|vaccine|"
    r"test substance|reference material|certified reference)", re.I
)


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def parse_size_from_sku(sku):
    m = re.match(r'^[\d]+\.[\d]+\.?(\d{4})$', sku)
    if not m:
        return ""
    last4 = m.group(1)
    if last4 == "0001":
        return ""
    if last4.startswith(("9", "6")):
        sized = int(last4[1:])
        if sized > 0:
            return f"{sized} L"
    sized = int(last4)
    if sized > 0:
        if sized >= 1000:
            return f"{sized} ml"
        else:
            return f"{sized} ml"
    return ""


def main():
    print("=== Cleanup Database ===\n")

    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb["Price List"]

    total = 0
    cat_changed = 0
    size_filled = 0
    folder_changed = 0

    for r in range(3, ws.max_row + 1):
        sku = str(ws.cell(r, 1).value or "").strip()
        if not sku:
            continue
        total += 1

        desc = str(ws.cell(r, 2).value or "")
        old_cat = str(ws.cell(r, 4).value or "")
        old_size = str(ws.cell(r, 5).value or "")
        old_folder = str(ws.cell(r, 9).value or "")
        old_main = old_folder.split("/")[0] if "/" in old_folder else ""

        # ---- Fix Category ----
        new_main = old_main
        new_cat = old_cat

        if sku.startswith("8."):
            new_main = "Chemicals"
        elif sku.startswith("1."):
            if EQUIPMENT_KEYWORDS.search(desc):
                new_main = "Equipment"
            else:
                new_main = "Chemicals"
        elif sku.startswith(("00-", "01.")):
            new_main = "Equipment"
        elif sku.startswith("02.") or sku.startswith(("03.", "04.")):
            if old_main not in CATEGORIES_AMAN:
                new_main = "Tools"
        elif sku.startswith("FT"):
            new_main = "Filter_Paper"

        if new_main != old_main and old_main in CATEGORIES_AMAN:
            new_main = old_main

        if new_main != old_main:
            cat_changed += 1
            folder_changed += 1
            if old_cat:
                new_cat = old_cat
            else:
                first_part = desc.split(",")[0].strip()
                words = re.findall(r"[A-Za-z]+", first_part)
                meaningful = [w for w in words if len(w) > 2 and w.lower() not in ("the", "for", "with", "and")]
                new_cat = meaningful[0].title() if meaningful else "Chemical"

            new_folder = f"{new_main}/{slugify(new_cat)}"
            ws.cell(r, 4, new_cat)
            ws.cell(r, 9, new_folder)
            if not old_size:
                size_from_sku = parse_size_from_sku(sku)
                if size_from_sku:
                    ws.cell(r, 5, size_from_sku)
                    size_filled += 1
        else:
            if not old_size and (sku.startswith(("1.", "8.")) or new_main == "Chemicals"):
                size_from_sku = parse_size_from_sku(sku)
                if size_from_sku:
                    ws.cell(r, 5, size_from_sku)
                    size_filled += 1

    wb.save(DB_PATH)
    print(f"Total SKU diproses: {total}")
    print(f"Kategori diubah: {cat_changed}")
    print(f"Size diisi dari SKU: {size_filled}")
    print(f"Folder diupdate: {folder_changed}")
    print("\nSelesai!")


if __name__ == "__main__":
    main()
