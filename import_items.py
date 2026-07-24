"""
Import SKU baru dari List Item Dinamika Lengkap.xlsx ke database.xlsx,
lalu isi otomatis field yang kosong (Brand, Category, Size, dll).
"""

import re
import os
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, "database.xlsx")
LIST_PATH = os.path.expanduser("~/Downloads/List Item Dinamika Lengkap.xlsx")

KNOWN_BRANDS = [
    "iwaki", "pyrex", "normax", "as one", "duran", "schott", "corning",
    "ika", "merck", "whatman", "eppendorf", "thermo", "nunc", "falcon",
    "sartorius", "brand", "witeg", "isolab", "boeco", "memmert", "binder",
    "wisd", "wisebath", "wisecure", "wisehot", "wisegen", "wise shaker",
    "wise vortex", "wise mixer", "dale", "rafa", "onemed", "gemet",
    "scilogex", "dlab", "delta lab",
]

MAIN_CATEGORIES = [
    "Glassware",
    "Tissue Culture",
    "Plastic Ware",
    "Porcelain Ware",
    "Tools",
    "Filter Paper",
    "Equipment",
    "Chemicals",
]

CATEGORY_KEYWORDS = [
    ("Tools",
     r"\b(?:tongs|spatula|scoop|forceps|pinset|scissors|"
     r"clamp|holder|stand|tripod|brush|rack\b|"
     r"loop|needle|knife|blade|file|cutter|hammer|"
     r"screwdriver|wrench|pipette ?filler|rubber ?bulb|"
     r"stopper.*(?:rubber|silicone)|septum|"
     r"thermometer|hydrometer|density.*cup|"
     r"rod.*stirrer|wire|gauze|trivet)\b"),

    ("Plastic Ware",
     r"\b(?:plastic|polypropylene|pp\b(?:\s|,|$)|pe\b(?:\s|,|$)|hdpe|ldpe|pvc|"
     r"polystyrene|petg|acrylic|nylon|polycarbonate|pmma|"
     r"microtube|microcentrifuge|eppendorf ?tube|"
     r"syringe|tip\b(?!s)|microtip|cuvette(?!.*glass)|"
     r"petri.*plastic|centrifuge.*(?:pp|plastic)|"
     r"conical.*(?:pp|plastic))\b"),

    ("Tissue Culture",
     r"\b(?:tissue ?culture|tc ?flask|cell ?culture|multiwell|"
     r"microplate|elisa|96 ?well|384 ?well|"
     r"culture ?dish|culture ?plate|serological|"
     r"cell ?scraper|culture ?flask)\b"),

    ("Equipment",
     r"\b(?:digital|electronic|analytical ?balance|ph ?meter|"
     r"spectrophotometer|oven|incubator|autoclave|centrifuge|"
     r"shaker|stirrer|hot ?plate|water ?bath|furnace|"
     r"bunsen ?burner|electronic ?burner|dispenser|pump|"
     r"magnetic ?stirrer|vortex|mixer|homogenizer|"
     r"rotary ?evaporator|laminar|cabinet|fume ?hood|"
     r"balance|conductivity ?meter|do ?meter|turbidity|"
     r"heater|block ?heater|dry ?bath|sonicator|"
     r"microscope|refractometer|polarimeter|melting ?point|"
     r"autotitrator|titrator|karl ?fischer|titrino)\b"),

    ("Filter Paper",
     r"\b(?:filter ?paper|membrane ?filter|whatman|"
     r"filter ?pad|filter ?disc|filter ?disk|filter ?circle|"
     r"filter ?sheet|filter ?roll|filter ?strip)\b"),

    ("Porcelain Ware",
     r"\b(?:porcelain|crucible|mortar|pestle|"
     r"evaporating ?dish|porcelain ?plate)\b"),

    ("Glassware",
     r"\b(?:beaker|flask|erlenmeyer|funnel|burette|buret|pipette|pipet|"
     r"condenser|condensor|test ?tube|petri|watch ?glass|"
     r"measuring ?cylinder|volumetric|desiccator|desicator|"
     r"crystallizing|boiling ?flask|distillation|soxhlet|soxhelet|"
     r"extractor|nessler|glass ?stopper|glass ?rod|"
     r"dropper|drying ?tube|filter ?holder|picnometer|pycnometer|"
     r"culture ?tube|ammonium|kjeldahl|sedimentation ?cone|"
     r"capillary|bottle|aspirator|iodine ?flask|automatic ?buret|"
     r"glass ?key|schelbach|liebig|allihn|west|davidson|"
     r"cuvette.*glass|slide ?glass|cover ?glass|"
     r"centrifuge.*glass|round ?bottom.*glass|"
     r"screw ?cap|cap\b(?:\s.*liner)?)\b"),

    ("Chemicals",
     r"\b(?:acid\b|base\b|buffer\b|solution|reagent|indicator|"
     r"dye\b|stain\b|solvent|alcohol|ethanol|methanol|acetone|"
     r"formaldehyde|sodium|potassium|calcium|magnesium|"
     r"chloride|sulfate|sulphate|nitrate|standard|certified|"
     r"reference ?material|reagent ?grade|"
     r"acs ?grade|hplc ?grade|gc ?grade|"
     r"ph ?buffer|conductivity ?standard|"
     r"titration|titrant|fixanal|volumetric ?standard)\b"),
]

SUBCATEGORY_KEYWORDS = [
    ("Beaker", r"\bbeaker\b"),
    ("Erlenmeyer", r"\berlenmeyer\b"),
    ("Volumetric Flask", r"\bvolumetric ?flask\b"),
    ("Boiling Flask", r"\bboiling ?flask\b"),
    ("Flask", r"\bflask\b"),
    ("Funnel", r"\bfunnel\b"),
    ("Burette", r"\bburette|buret\b"),
    ("Pipette", r"\bpipette|pipet\b"),
    ("Measuring Cylinder", r"\bmeasuring ?cylinder|graduated ?cylinder\b"),
    ("Test Tube", r"\btest ?tube\b"),
    ("Culture Tube", r"\bculture ?tube\b"),
    ("Centrifuge Tube", r"\bcentrifuge ?tube|conical ?tube\b"),
    ("Petri Dish", r"\bpetri ?dish|petri\b"),
    ("Watch Glass", r"\bwatch ?glass\b"),
    ("Condenser", r"\bcondenser|condensor\b"),
    ("Distillation Apparatus", r"\bdistillation\b"),
    ("Extractor", r"\bextractor\b"),
    ("Desiccator", r"\bdesiccator|desicator\b"),
    ("Bottle", r"\bbottle\b"),
    ("Dropper", r"\bdropper\b"),
    ("Glass Stopper", r"\bglass ?stopper\b"),
    ("Plastic Stopper", r"\bplastic ?stopper|stopper\b"),
    ("Screw Cap", r"\bscrew ?cap|cap\b"),
    ("Glass Rod", r"\bglass ?rod|stirring ?rod|stirrer\b"),
    ("Drying Tube", r"\bdrying ?tube\b"),
    ("Filter Holder", r"\bfilter ?holder\b"),
    ("Picnometer", r"\bpicnometer|pycnometer\b"),
    ("Sampling Tube", r"\bsampling ?tube\b"),
    ("Sedimentation Cone", r"\bsedimentation ?cone\b"),
    ("Crystallizing Dish", r"\bcrystallizing ?dish\b"),
    ("Microscope Slide", r"\bmicroscope ?slide|slide ?glass\b"),
    ("Cover Glass", r"\bcover ?glass\b"),
    ("Nessler Tube", r"\bnessler\b"),
    ("Digestion Tube", r"\bdigestion ?tube\b"),
    ("Bunsen Burner", r"\bbunsen ?burner\b"),
    ("Tongs", r"\btongs\b"),
    ("Spatula", r"\bspatula\b"),
    ("Scissors", r"\bscissors\b"),
    ("Forceps", r"\bforceps|pinset\b"),
    ("Clamp", r"\bclamp\b"),
    ("Holder", r"\bholder\b"),
    ("Stand", r"\bstand\b"),
    ("Rack", r"\brack\b"),
    ("Crucible", r"\bcrucible\b"),
    ("Mortar", r"\bmortar\b"),
    ("Pestle", r"\bpestle\b"),
    ("Porcelain Plate", r"\bporcelain ?plate\b"),
    ("Filter Paper", r"\bfilter ?paper\b"),
    ("Membrane Filter", r"\bmembrane ?filter\b"),
    ("Syringe", r"\bsyringe\b"),
    ("Needle", r"\bneedle\b"),
    ("Tips", r"\btip\b"),
    ("Cuvette", r"\bcuvette\b"),
    ("Microtube", r"\bmicrotube|micro ?tube\b"),
    ("Thermometer", r"\bthermometer\b"),
    ("Hydrometer", r"\bhydrometer\b"),
    ("Brush", r"\bbrush\b"),
]

VISION_KEYWORDS_TEMPLATE = {
    "Glassware": "clear glass laboratory {subcategory} equipment",
    "Tissue Culture": "sterile plastic laboratory {subcategory} for cell culture",
    "Plastic Ware": "plastic laboratory {subcategory} disposable equipment",
    "Porcelain Ware": "white porcelain laboratory {subcategory} heat resistant",
    "Tools": "laboratory {subcategory} tool stainless steel",
    "Filter Paper": "laboratory grade {subcategory} for filtration",
    "Equipment": "laboratory {subcategory} equipment",
    "Chemicals": "laboratory grade {subcategory} chemical reagent",
}


def load_database_skus(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Price List"]
    skus = set()
    for r in range(3, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        if sku:
            skus.add(str(sku).strip())
    return skus


def load_list_items(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    items = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        if not sku or not desc:
            continue
        items.append((str(sku).strip(), str(desc).strip()))
    return items


def detect_brand(text):
    text_lower = text.lower()
    for brand in KNOWN_BRANDS:
        if brand in text_lower:
            return brand.title()
    return ""


def extract_size(text):
    patterns = [
        # volume: 10 ml, 100 mL, 0.5 L, 1000 μL, etc (case insensitive)
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*(ml|l|μ[lL]|ul|µ[lL]|liter)\b",
        # dimension: 10 x 75 mm, D 50 x H 15 mm
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*[x×]\s*\d+(?:[\.\,]\d+)?\s*(mm|cm|m|inch|in)\b",
        # dimension with prefix: D 11, H 130, L 200, W 98, ID 150, OD 60
        r"(?i)(?:d|h|l|w|dt|id|od)\s*[:=]?\s*(\d+(?:[\.\,]\d+)?)\s*(mm|cm|m|inch|in)\b",
        # standalone number + unit (longer units first to avoid partial match)
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*(mm|cm|inch|in)\b",
        # number + unit in parentheses
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*\((mm|cm|m|inch|in|ml|l)\)",
        # capacity/weight (case insensitive)
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*(g|mg|kg|lb|oz)\b(?!\w)",
        # Dia prefix: Dia 90mm, Dia 90 mm
        r"(?i)dia\s*\.?\s*(\d+(?:[\.\,]\d+)?)\s*(mm|cm|m|inch|in)\b",
# percentage: 0-100%
        r"(?i)(\d+(?:[\.\,]\d+)?)\s*(%)",
    ]
    found = []
    seen = set()
    for pat in patterns:
        matches = re.findall(pat, text)
        for m in matches:
            val = m[0].replace(",", ".")
            try:
                val_f = float(val)
                if val_f == int(val_f):
                    val = str(int(val_f))
            except ValueError:
                val = m[0]
            unit = m[1]
            key = f"{val}{unit}".lower()
            if key not in seen:
                seen.add(key)
                found.append(f"{val} {unit}")
    # TS/GCMI pattern (non-standard format: unit comes first)
    m_ts = re.search(r"(?i)TS\s*(\d+(?:/\d+)?)", text)
    if m_ts:
        key = f"ts{m_ts.group(1)}".lower()
        if key not in seen:
            seen.add(key)
            found.append(f"TS {m_ts.group(1)}")
    m_gcmi = re.search(r"(?i)GCMI\s*(\d+(?:-\d+)?)", text)
    if m_gcmi:
        key = f"gcmi{m_gcmi.group(1)}".lower()
        if key not in seen:
            seen.add(key)
            found.append(f"GCMI {m_gcmi.group(1)}")
    m_model = re.search(r"(?i)\b(FK|cm2)\s*(\d+)", text)
    if m_model:
        key = f"{m_model.group(1)}{m_model.group(2)}".lower()
        if key not in seen:
            seen.add(key)
            unit = m_model.group(1).upper()
            found.append(f"{m_model.group(2)} {unit}")
    if found:
        return ", ".join(found[:3])
    return ""


def detect_main_category(text):
    text_lower = text.lower()
    for cat, pattern in CATEGORY_KEYWORDS:
        if re.search(pattern, text_lower):
            return cat
    return "Equipment"


def detect_subcategory(text, main_cat):
    text_lower = text.lower()
    for subcat, pattern in SUBCATEGORY_KEYWORDS:
        if re.search(pattern, text_lower):
            return subcat
    first_part = text.split(",")[0].strip()
    first_part = re.sub(r"\(.*?\)", "", first_part).strip()
    words = re.findall(r"[A-Za-z]+", first_part)
    meaningful = [w for w in words if len(w) > 2 and w.lower() not in ("the", "for", "with", "and")]
    if meaningful:
        return meaningful[0].title()
    return main_cat


def generate_ocr_keywords(desc, brand, size, subcat):
    parts = [desc.lower()]
    if brand and brand.lower() not in desc.lower():
        parts.append(brand.lower())
    if size and size.lower() not in desc.lower():
        parts.append(size.lower())
    if subcat.lower() not in desc.lower():
        parts.append(subcat.lower())
    return ", ".join(parts)


def generate_vision_keywords(main_cat, subcat):
    template = VISION_KEYWORDS_TEMPLATE.get(main_cat, "laboratory {subcategory}")
    return template.format(subcategory=subcat.lower())


def slugify(text):
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9\s]", "", text)
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def item_to_row(sku, desc):
    brand = detect_brand(desc)
    main_cat = detect_main_category(desc)
    subcat = detect_subcategory(desc, main_cat)
    size = extract_size(desc)
    ocr_kw = generate_ocr_keywords(desc, brand, size, subcat)
    vision_kw = generate_vision_keywords(main_cat, subcat)
    template = subcat
    output_folder = f"{main_cat.replace(' ', '_')}/{slugify(subcat)}"
    active = "=TRUE()"
    return [sku, desc, brand, subcat, size, ocr_kw, vision_kw, template, output_folder, active]


def main():
    print("=== Import SKU Baru ===\n")

    if not os.path.exists(LIST_PATH):
        print(f"ERROR: File tidak ditemukan: {LIST_PATH}")
        return

    existing_skus = load_database_skus(DATABASE_PATH)
    print(f"SKU yang sudah ada di database: {len(existing_skus)}")

    items = load_list_items(LIST_PATH)
    print(f"Total item di list: {len(items)}")

    new_items = [(s, d) for s, d in items if s not in existing_skus]
    print(f"SKU baru yang akan ditambahkan: {len(new_items)}")

    if not new_items:
        print("Tidak ada SKU baru. Selesai.")
        return

    wb = openpyxl.load_workbook(DATABASE_PATH)
    ws = wb["Price List"]
    next_row = ws.max_row + 1

    cat_counts = {}
    added = 0
    for sku, desc in new_items:
        row = item_to_row(sku, desc)
        for c in range(1, len(row) + 1):
            ws.cell(next_row, c, row[c - 1])
        next_row += 1
        added += 1

        main_cat = row[3]
        cat_counts[main_cat] = cat_counts.get(main_cat, 0) + 1

        if added % 1000 == 0:
            print(f"  Progress: {added}/{len(new_items)}")

    wb.save(DATABASE_PATH)
    print(f"\nBerhasil menambahkan {added} SKU baru ke database.xlsx!")

    print("\nBreakdown per kategori:")
    for cat in MAIN_CATEGORIES:
        count = cat_counts.get(cat, 0)
        pct = count / added * 100 if added else 0
        print(f"  {cat:20s}: {count:6d} ({pct:5.1f}%)")
    for cat, count in sorted(cat_counts.items()):
        if cat not in MAIN_CATEGORIES:
            print(f"  {cat:20s}: {count:6d}")

    print(f"\nTotal row di database sekarang: {next_row - 1}")


if __name__ == "__main__":
    main()
